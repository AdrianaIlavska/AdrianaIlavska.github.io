#!/usr/bin/env python3
"""Update Pocet_praci values in index.js and index_EN.js from an XLSX file."""

from __future__ import annotations

import argparse
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with `python3 -m pip install openpyxl`."
    ) from exc


DEFAULT_XLSX = "vytizenost_2026_jar.xlsx"
DEFAULT_JS_FILES = ("index.js", "index_EN.js")
NAME_COLUMNS = ("VEDOUCI", "JMENO", "JMÉNO", "NAME")
VALUE_COLUMNS = ("SUM_POCET", "POCET_PRACI")

ROW_PATTERN = re.compile(
    r"(?P<prefix>Vedouci:\s*(['\"])(?P<name>(?:\\.|(?!\2).)*)\2"
    r"(?P<middle>.*?Pocet_praci:\s*))"
    r"(?P<count>-?\d+(?:\.\d+)?)",
    re.DOTALL,
)


def normalize_header(value: object) -> str:
    return str(value or "").strip().upper()


def normalize_name(value: object) -> str:
    text = str(value or "").casefold().strip()
    text = re.sub(r"\s+", " ", text)
    return text.rstrip(". ")


def format_count(value: object) -> str:
    if value is None or value == "":
        raise ValueError("empty count")

    try:
        number = Decimal(str(value).strip())
    except InvalidOperation as exc:
        raise ValueError(f"invalid count {value!r}") from exc

    if number != number.to_integral_value():
        raise ValueError(f"count is not a whole number: {value!r}")

    return str(int(number))


def find_column(headers: list[object], candidates: tuple[str, ...]) -> int | None:
    normalized = [normalize_header(header) for header in headers]
    for candidate in candidates:
        if candidate in normalized:
            return normalized.index(candidate)
    return None


def read_counts(xlsx_path: Path, sheet_name: str | None) -> dict[str, str]:
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

    rows = sheet.iter_rows(values_only=True)
    try:
        headers = list(next(rows))
    except StopIteration as exc:
        raise ValueError(f"{xlsx_path} is empty") from exc

    name_col = find_column(headers, NAME_COLUMNS)
    value_col = find_column(headers, VALUE_COLUMNS)

    if name_col is None:
        raise ValueError(
            "Could not find a supervisor-name column. Expected one of: "
            + ", ".join(NAME_COLUMNS)
        )
    if value_col is None:
        raise ValueError(
            "Could not find a count column. Expected SUM_POCET, or Pocet_praci as fallback."
        )

    counts: dict[str, str] = {}
    for row_number, row in enumerate(rows, start=2):
        if name_col >= len(row) or value_col >= len(row):
            continue

        raw_name = row[name_col]
        if raw_name is None or str(raw_name).strip() == "":
            continue

        name = normalize_name(raw_name)
        count = format_count(row[value_col])
        if name in counts:
            raise ValueError(f"Duplicate supervisor name in XLSX on row {row_number}: {raw_name}")
        counts[name] = count

    return counts


def update_js_file(js_path: Path, counts: dict[str, str], dry_run: bool) -> tuple[int, list[str]]:
    text = js_path.read_text(encoding="utf-8")
    updated = 0
    unmatched: list[str] = []

    def replace(match: re.Match[str]) -> str:
        nonlocal updated
        display_name = match.group("name")
        key = normalize_name(display_name)
        if key not in counts:
            unmatched.append(display_name)
            return match.group(0)

        new_count = counts[key]
        if match.group("count") != new_count:
            updated += 1
        return f"{match.group('prefix')}{new_count}"

    new_text = ROW_PATTERN.sub(replace, text)
    if not dry_run and new_text != text:
        js_path.write_text(new_text, encoding="utf-8")

    return updated, unmatched


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Update Pocet_praci values in JS files from an XLSX column."
    )
    parser.add_argument("xlsx", nargs="?", default=DEFAULT_XLSX, help=f"default: {DEFAULT_XLSX}")
    parser.add_argument(
        "--js",
        nargs="+",
        default=list(DEFAULT_JS_FILES),
        help="JS files to update. Default: index.js index_EN.js",
    )
    parser.add_argument("--sheet", help="Worksheet name. Default: first sheet.")
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing files.")
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Exit with an error if any JS supervisor is missing from the XLSX.",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"XLSX file not found: {xlsx_path}", file=sys.stderr)
        return 1

    counts = read_counts(xlsx_path, args.sheet)
    print(f"Loaded {len(counts)} supervisor counts from {xlsx_path}.")

    any_unmatched = False
    for js_file in args.js:
        js_path = Path(js_file)
        if not js_path.exists():
            print(f"JS file not found: {js_path}", file=sys.stderr)
            return 1

        changed, unmatched = update_js_file(js_path, counts, args.dry_run)
        action = "Would update" if args.dry_run else "Updated"
        print(f"{action} {changed} Pocet_praci value(s) in {js_path}.")

        if unmatched:
            any_unmatched = True
            print(f"  No XLSX match for {len(unmatched)} supervisor(s):")
            for name in unmatched:
                print(f"  - {name}")

    if args.strict and any_unmatched:
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
